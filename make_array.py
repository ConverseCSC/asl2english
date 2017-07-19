import os

import openpyxl

import urllib.request

import statistics


# ASL Glossary spreadsheet organization: 
#  	Column B = Sign names starting at B3, separated by different number empty spaces. 
#	Column K = Movie files in center, occasionally more than 1 present

# Goal: must match up sign names with movie files, keep multiple movie files using 
#			file1, file2, file3, etc naming convention

def video_search():
	
	book = openpyxl.load_workbook('ASL_Glossary.xlsx')

	ws = book.active
		
	video_links = []
		
	for col in ws.iter_cols(min_row=3, max_row=10, min_col=2, max_col=2):
		
		for cell in col:
			
			if cell.value != None and cell.value != " ":
				next_row = cell.row + 1
				next_cell = ws['B' + str(next_row)]
		
				empty_block = [] # will contain list of cells between signs	
					
				while next_cell.value == None or next_cell.value == " ":
					
					empty_block.append(next_cell)
					next_row = int(next_row) + 1
					next_cell = ws['B' + str(next_row)]
						
						
			# Iterate over empty_block to check corresponding cells for movie files
		
				videos = []
		
				for empty_cell in empty_block:
					
					empty_row = empty_cell.row
						
					video = ws['K' + str(empty_row)]
						
					if video.value != None and video.value != " " and video.value != '------------':
						
						url_obj = video.hyperlink
						url = url_obj.target
						videos.append(url)
				video_links.append(videos)	
		
	return video_links
		
def name_search():
	
	book = openpyxl.load_workbook('ASL_Glossary.xlsx')

	ws = book.active
		
	sign_names = []
		
	for col in ws.iter_cols(min_row=3, max_row=10, min_col=2, max_col=2):
		
		for cell in col:
			
			if cell.value != None and cell.value != " ":
				sign_name = cell.value
				sign_names.append(sign_name)
		
	return sign_names
		
		
def file_naming():
	
	file_names_array = []
		
	for n, v in zip(name_search(), video_search()):
		if len(v) > 1:
			counter = 0
			names = []
			for i in v:
				counter = int(counter) + 1
				file_name = str(n).lower() + str(counter)
				names.append(file_name + '.mp4')
			file_names_array.append(names)
						
		else: 
			file_name = str(n).lower() 
			file_names_array.append(file_name + '.mp4')
		
	sign_array = make_array()
		
	for f, s in zip(file_names_array, sign_array):
		s['video'] = f

	return sign_array
		
		

def make_array():
	
	book = openpyxl.load_workbook('ASL_Glossary.xlsx')

	ws = book.active
		
	signs = []
		
	for col in ws.iter_cols(min_row=3, max_row=10, min_col=2, max_col=2):
		
		for cell in col:
				
			if cell.value != None and cell.value != " ":
				sign_name = cell.value
				next_row = cell.row + 1
				next_cell = ws['B' + str(next_row)]
		
				empty_block = [] # will contain list of cells between signs	
					
				while next_cell.value == None or next_cell.value == " ":
					
					empty_block.append(next_cell)
					next_row = int(next_row) + 1
					next_cell = ws['B' + str(next_row)]
						
					
		# Iterate over empty_block to check corresponding cells for movie files
		
				ds_handshapes = []
				de_handshapes = []
				nds_handshapes = []
				nde_handshapes = []
				videos = []
		
				for empty_cell in empty_block:
					
					empty_row = empty_cell.row
						
					# Hand shapes (dominant and non dominant)
					
					d_start = ws['F' + str(empty_row)]
					d_end = ws['H' + str(empty_row)]
					nd_start = ws['G' + str(empty_row)]
					nd_end = ws['I' + str(empty_row)]
						
					
						
						
						# Dominant starting hand shape
							
					if d_start.value != None and d_start.value != " " and d_start.value != "------------":
							
						ds_handshapes.append(d_start.value)
					
					if d_end.value != None and d_end.value != " " and d_end.value != "------------":
							
						de_handshapes.append(d_end.value)
						
					if nd_start.value != None and nd_start.value != " " and nd_start.value != "------------":
						
						nds_handshapes.append(nd_start.value)
							
						# Non Dominant ending hand shape
						
					if nd_end.value != None and nd_end.value != " " and nd_end.value != "------------":
							
						nde_handshapes.append(nd_end.value)
						
							
							
					
				# Finds the most commonly used hand shape for that sign	
				if ds_handshapes == []:	
					ds_handshape = None
				else: 
					try:
						ds_handshape = statistics.mode(ds_handshapes)
					except statistics.StatisticsError:
						ds_handshape = ds_handshapes
				if nds_handshapes == []:
					nds_handshape = None
					hand_number = 1
				else: 
					try:
						nds_handshape = statistics.mode(nds_handshapes)
					except statistics.StatisticsError:
						nds_handshape = nds_handshapes
					hand_number = 2
				if de_handshapes == []:
					de_handshape = None
				else:
					try:
						de_handshape = statistics.mode(de_handshapes)
					except statistics.StatisticsError:
						de_handshape = de_handshapes
				if nde_handshapes == []:
					nde_handshape = None
				else:
					try:
						nde_handshape = statistics.mode(nde_handshapes)
					except statistics.StatisticsError:
						nde_handshape = nde_handshapes
					
				dominant_handshapes = [ds_handshape, de_handshape]
				
				nondominant_handshapes = [nds_handshape, nde_handshape]
					
				handshapes = [dominant_handshapes, nondominant_handshapes]
					
					
					
				sign = {'sign': sign_name, 'hands': hand_number, 
				'handshape': handshapes, 'position': [], 'palmface': '', 
				'motion': {'type': '', 'dir': []}}
					
				signs.append(sign)			
					
	return signs
					
# Reminder to remove max_row = 20 when doing full test

file_naming()


