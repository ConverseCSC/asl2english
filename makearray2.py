import os

import openpyxl

import urllib.request


def variant_search():
	
	book = openpyxl.load_workbook('ASL_Glossary.xlsx')

	ws = book.active
		
	video_links = []
	video_block = []
	variants = []
	
	for col in ws.iter_cols(min_row=3, max_row=13232, min_col=11, max_col=11):
		
		for cell in col:
			if cell.value == '============':
				next_row = cell.row + 1
				next_cell = ws['K' + str(next_row)]

				empty_block = [] # will contain list of cells between signs
				
				while next_cell.value != '============':
					empty_block.append(next_cell)
					next_row = int(next_row) + 1
					next_cell = ws['K' + str(next_row)]
						
				videos = []
				
				for empty_cell in empty_block:
					if empty_cell.value == '------------':
						if empty_block not in video_block:
							video_block.append(empty_block)
				
	for block in video_block:
		for video_cell in block:
			if video_cell.value != '------------' and video_cell.value != " " and video_cell.value != None:
				vnext_row = video_cell.row + 1
				vprev_row = video_cell.row - 1
				vnext_cell = ws['K' + str(vnext_row)]
				vprev_cell = ws['K' + str(vprev_row)]
			
				variant_block = []
	
				
				while vprev_cell.value != '------------' and vprev_cell.value != '============':
					variant_block.insert(0, vprev_cell)
					vprev_row = int(vprev_row) - 1
					vprev_cell = ws['K' + str(vprev_row)]
					
				variant_block.append(video_cell)
				
				while vnext_cell.value != '------------' and vnext_cell.value != '============':
					variant_block.append(vnext_cell)
					vnext_row = int(vnext_row) + 1
					vnext_cell = ws['K' + str(vnext_row)]
				
				variants.append(variant_block)	
	return variants
						
					
def name_search():
	
	book = openpyxl.load_workbook('ASL_Glossary.xlsx')

	ws = book.active
		
	sign_names = []
	
	
	variants = variant_search()
	
	for sign_var in variants:
		sign_var_row = sign_var[0].row - 1
		name_cell = ws['B' + str(sign_var_row)]
		variant_name_cell = ws['E' + str(sign_var[0].row)]
		
		if name_cell.value != None and name_cell.value != " ":
			sign_names.append(name_cell.value + ': ' + variant_name_cell.value)
		else:
			while name_cell.value == None or name_cell.value == " ":
				sign_var_row = int(sign_var_row) - 1
				name_cell = ws['B' + str(sign_var_row)]
			sign_names.append(name_cell.value + ': ' + variant_name_cell.value)

	return sign_names

def video_name_search():
	
	book = openpyxl.load_workbook('ASL_Glossary.xlsx')

	ws = book.active
		
	video_names = []
	
	
	variants = variant_search()

	
	for sign_var in variants:
		sign_var_row = sign_var[0].row - 1
		name_cell = ws['B' + str(sign_var_row)]
		variant_name_cell = ws['E' + str(sign_var[0].row)]
		
		if name_cell.value != None and name_cell.value != " ":
			video_names.append(name_cell.value)
		else:
			while name_cell.value == None or name_cell.value == " ":
				sign_var_row = int(sign_var_row) - 1
				name_cell = ws['B' + str(sign_var_row)]
			video_names.append(name_cell.value)

	return video_names
		
		
def file_naming():
	
	file_names_array = []
	counter = 0
	names = video_name_search()
		
	for i in range(0, len(names)):
		if i == len(names) - 1:
			break 
		elif names[i] == names[int(i) + 1]:
			counter = int(counter) + 1
			file_name = names[i].lower() + str(counter)
			file_names_array.append(file_name + '.mp4')
		elif names[i] == names[int(i) - 1]:
			counter = int(counter) + 1
			file_name = names[i].lower() + str(counter)
			file_names_array.append(file_name + '.mp4')
			counter = 0
				
		
	sign_array = make_array()
	
	names = name_search()
		
	for f, s in zip(file_names_array, sign_array):
		s['video'] = f
		
	for n, a in zip(names, sign_array):
		a['sign'] = n

	return sign_array

	
		

def make_array():
	book = openpyxl.load_workbook('ASL_Glossary.xlsx')

	ws = book.active

	variants = variant_search()
	
	sign_names = name_search()
	
	signs = []
	
	for var_block in variants:		
		ds_handshapes = []
		nds_handshapes = []
		de_handshapes = []
		nde_handshapes = []
		videos = []
		
		for var_cell in var_block:
					
			var_row = var_cell.row
						
			# Hand shapes (dominant and non dominant)
					
			d_start = ws['F' + str(var_row)]
			nd_start = ws['G' + str(var_row)]
			d_end = ws['H' + str(var_row)]
			nd_end = ws['I' + str(var_row)]
						
						
				# Dominant starting hand shape
					
			if d_start.value != None and d_start.value != " ":
					
				ds_handshapes.append(str(d_start.value))
					
				# Non Dominant starting hand shape
			
			if nd_start.value != None and nd_start.value != " ":
					
				nds_handshapes.append(str(nd_start.value))
					
				# Dominant ending hand shape
			
			if d_end.value != None and d_end.value != " ":
					
				de_handshapes.append(str(d_end.value))
					
				# Non Dominant ending hand shape
				
			if nd_end.value != None and nd_end.value != " ":
					
				nde_handshapes.append(str(nd_end.value))
						
# 			Finds the most commonly used hand shape for that sign
		d_handshape = []
		nd_handshape = []	
		if ds_handshapes == [] and de_handshapes == []:	
			d_handshape = None
		elif ds_handshapes == []:
			ds_handshapes = None
			d_handshape.append(ds_handshapes)
		elif len(ds_handshapes) == 1:
			d_handshape.append(ds_handshapes[0])
		else: 
			d_handshape.append(ds_handshapes)
		
		if ds_handshapes == [] and de_handshapes == []:
			d_handshape = None
		elif len(de_handshapes) == 1:
			d_handshape.append(de_handshapes[0])
		elif de_handshapes == []:
			de_handshapes = None
			d_handshape.append(de_handshapes)
		else:	
			d_handshape.append(de_handshapes)
			
		
		if nds_handshapes == [] and nde_handshapes == []:	
			nd_handshape = None
			hand_number = 1
		elif nds_handshapes == []:
			nds_handshapes = None
			nd_handshape.append(nds_handshapes)
		elif len(nds_handshapes) == 1:
			nd_handshape.append(nds_handshapes[0])
			hand_number = 2
		else: 
			nd_handshape.append(nds_handshapes)
			hand_number = 2
		
		if nds_handshapes == [] and nde_handshapes == []:
			nd_handshape = None
		elif len(nde_handshapes) == 1:
			nd_handshape.append(nde_handshapes[0])
			hand_number = 2
		elif nde_handshapes == []:
			nde_handshapes = None
			nd_handshape.append(nde_handshapes)
		else:	
			nd_handshape.append(nde_handshapes)
			hand_number = 2
			
		
	
			
		if nd_handshape == None:
			handshapes = [d_handshape]
		else:
			handshapes = [d_handshape, nd_handshape]
			

			
		sign = {'hands': hand_number, 
		'handshape': handshapes, 'position': [], 'palmface': '', 
		'motion': {'type': '', 'dir': []}}
			
		signs.append(sign)			
					
	return signs
# # 					
# # Reminder to remove max_row = 20 when doing full test




